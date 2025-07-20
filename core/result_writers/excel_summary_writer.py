# core/writers/excel_writer.py

import os
from typing import Optional

from openpyxl import Workbook, load_workbook
from openpyxl.styles import PatternFill
from openpyxl.utils import get_column_letter


class ExcelSummaryWriter:
    def __init__(self, summary_path="data/output.xlsx"):
        self.path = summary_path
        self._closed = False
        
        # Create directory if it doesn't exist
        os.makedirs(os.path.dirname(self.path), exist_ok=True)
        
        if not os.path.exists(self.path):
            # Create a new Excel file if template doesn't exist
            print(f"📝 Creating new Excel file at '{self.path}'")
            self.wb = Workbook()
            # Remove default sheet and create a summary sheet
            self.wb.remove(self.wb.active)
            self.wb.create_sheet("Summary")
        else:
            try:
                self.wb = load_workbook(self.path)
            except Exception as e:
                print(f"⚠️ Could not load existing Excel file: {e}")
                print("📝 Creating new Excel file...")
                self.wb = Workbook()
                self.wb.remove(self.wb.active)
                self.wb.create_sheet("Summary")
        
        self.RULE_ROW_MAPPING = [
            0,                  # MarginRule
            (1, 2),             # CoverPageTitleRule + CoverPageTableRule → merge
            3,                  # TableOfContentsRule
            4,                  # StrictHeading1Rule
            5,                  # StrictHeading2Rule
            6,                  # CombinedStyleRule
            7,                  # ImageRightOfTextRule
            8,                  # FootnoteOnHabitatRule
            9,                  # FooterRule
            10,                 # CombinedMultilevelListRule
            11,                 # PageBreakBeforeHeadingRule
            12                  # ReferencesHangingIndentRule
        ]

    def _convert_result_to_excel_marks(self, result: dict) -> list[int]:
        """
        Convert 13-rule result into 12-row mark list for Excel, with merge logic applied.
        """
        results = result["results"]
        marks = [None] * 12
        needs_review = [False] * 12

        for i, rule_map in enumerate(self.RULE_ROW_MAPPING):
            if isinstance(rule_map, tuple):
                combined_mark = sum(results[idx]["mark"] for idx in rule_map)
                marks[i] = combined_mark

                combined_review = any(results[idx].get("needs_review", False) for idx in rule_map)
                needs_review[i] = combined_review
            else:
                marks[i] = results[rule_map]["mark"]
                needs_review[i] = results[rule_map].get("needs_review", False)

        return marks, needs_review

    def write(self, zid: str, result: dict):
        """
        Write the result for the given zid into the correct column of all sheets.
        """
        if self._closed or not hasattr(self, 'wb') or self.wb is None:
            print(f"⚠️ Cannot write to Excel for {zid}: workbook is closed")
            return
            
        try:
            WARNING_FILL = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")
            marks, needs_review_list = self._convert_result_to_excel_marks(result)

            for sheet in self.wb.worksheets:
                row_2 = list(sheet.iter_rows(min_row=2, max_row=2, values_only=True))[0]
                for col_idx in range(6, sheet.max_column + 1):
                    zid_value = row_2[col_idx - 1]
                    if isinstance(zid_value, str) and zid.lower() in zid_value.lower():
                        for i, (mark, needs_review) in enumerate(zip(marks, needs_review_list)):
                            cell = sheet.cell(row=4 + i, column=col_idx)
                            if needs_review:
                                cell.value = f"{mark} ⚠️" if mark is not None else "⚠️"
                                cell.fill = WARNING_FILL
                            else:
                                cell.value = mark
        except Exception as e:
            print(f"❌ Error writing to Excel for {zid}: {e}")

    def save(self, output_path: Optional[str] = None):
        """Save the workbook safely"""
        if self._closed:
            print("⚠️ Cannot save: workbook is already closed")
            return
            
        try:
            if hasattr(self, 'wb') and self.wb is not None:
                save_path = output_path or self.path
                self.wb.save(save_path)
                print(f"✅ Excel file saved successfully to {save_path}")
        except Exception as e:
            print(f"❌ Failed to save Excel file: {e}")

    def close(self):
        """Safely close the workbook"""
        if self._closed:
            return
            
        try:
            if hasattr(self, 'wb') and self.wb is not None:
                self.save()
                self.wb.close()
                self.wb = None
            self._closed = True
        except Exception as e:
            print(f"❌ Error closing Excel file: {e}")

    def __del__(self):
        """Destructor - ensure file is saved before object is destroyed"""
        if not self._closed:
            try:
                if hasattr(self, 'wb') and self.wb is not None:
                    self.save()
            except Exception as e:
                print(f"❌ Failed to save Excel file in destructor: {e}")
